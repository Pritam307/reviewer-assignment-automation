"""
Reviewer Assignment Automation Script
=====================================
This script:
1. Connects to MySQL database
2. Runs the reviewer assignment query
3. Generates an Excel report
4. Sends email notifications to reviewers

Author: [Your Name]
Date: January 2026
"""

import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
import pandas as pd
import mysql.connector
from mysql.connector import Error
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
load_dotenv()
# =============================================================================
# CONFIGURATION
# =============================================================================

# Database configuration (from environment variables)
DB_CONFIG = {
    'host': os.environ.get('DB_HOST'),
    'port': int(os.environ.get('DB_PORT', 3306)),
    'database': os.environ.get('DB_NAME'),
    'user': os.environ.get('DB_USER'),
    'password': os.environ.get('DB_PASSWORD')
}

# Email configuration (from environment variables)
# EMAIL_CONFIG = {
#     'smtp_host': os.environ.get('SMTP_HOST'),
#     'smtp_port': int(os.environ.get('SMTP_PORT', 587)),
#     'smtp_user': os.environ.get('SMTP_USER'),
#     'smtp_password': os.environ.get('SMTP_PASSWORD'),
#     'from_email': os.environ.get('EMAIL_FROM')
# }
# Email configuration (Gmail)
EMAIL_CONFIG = {
    'smtp_host': 'smtp.gmail.com',
    'smtp_port': 587,
    'smtp_user': os.environ.get('GMAIL_USER', 'abc123@gmail.com'),
    'smtp_password': os.environ.get('GMAIL_APP_PASSWORD'),
    'from_email': os.environ.get('GMAIL_USER', 'abc123@gmail.com')
}


# Reviewer email mapping
# REVIEWER_EMAILS = {
#     'Reviewer 1': os.environ.get('REVIEWER_1_EMAIL'),
#     'Reviewer 2': os.environ.get('REVIEWER_2_EMAIL'),
#     'Reviewer 3': os.environ.get('REVIEWER_3_EMAIL'),
#     'Reviewer 4': os.environ.get('REVIEWER_4_EMAIL'),
#     'Reviewer 5': os.environ.get('REVIEWER_5_EMAIL')
# }


# =============================================================================
# SQL QUERY
# =============================================================================

REVIEWER_ASSIGNMENT_QUERY = """
WITH 
-- Step 1: Get all article-product pairs for the given criteria
base_data AS (
    SELECT 
        tpa.ArticleId,
        tp.Id AS ProductId,
        tp.ProductName
    FROM tbl_product_articles tpa
    INNER JOIN tbl_products tp ON tpa.ProductId = tp.Id
    INNER JOIN tbl_articles ta ON tpa.ArticleId = ta.ID
    WHERE tp.ClientID = 3
      AND tpa.DownloadedDate BETWEEN %s AND %s
),

-- Step 2: Count occurrences per product
product_counts AS (
    SELECT 
        ProductId,
        ProductName,
        COUNT(*) AS occurrence_count
    FROM base_data
    GROUP BY ProductId, ProductName
),

-- Step 3: Calculate total rows
total_rows AS (
    SELECT SUM(occurrence_count) AS total_count FROM product_counts
),

-- Step 4: Rank products by occurrence
product_ranked AS (
    SELECT 
        pc.ProductId,
        pc.ProductName,
        pc.occurrence_count,
        ROW_NUMBER() OVER (ORDER BY pc.occurrence_count DESC, pc.ProductName ASC) AS product_rank,
        SUM(pc.occurrence_count) OVER (ORDER BY pc.occurrence_count DESC, pc.ProductName ASC) AS cumulative_count,
        tr.total_count
    FROM product_counts pc
    CROSS JOIN total_rows tr
),

-- Step 5: Determine the split point
split_point AS (
    SELECT MIN(product_rank) AS split_rank
    FROM product_ranked
    WHERE cumulative_count >= total_count / 2
),

-- Step 6: Classify high/low weightage
product_with_high_low AS (
    SELECT 
        pr.ProductId,
        pr.ProductName,
        pr.occurrence_count,
        pr.product_rank,
        CASE WHEN pr.product_rank <= sp.split_rank THEN 1 ELSE 0 END AS is_high_weightage
    FROM product_ranked pr
    CROSS JOIN split_point sp
),

-- Step 7: Assign reviewers to products
product_final_assignment AS (
    SELECT 
        ProductId,
        ProductName,
        occurrence_count,
        product_rank,
        is_high_weightage,
        CASE 
            WHEN is_high_weightage = 1 THEN 
                CASE WHEN (ROW_NUMBER() OVER (PARTITION BY is_high_weightage ORDER BY occurrence_count DESC, ProductName ASC) - 1) % 2 = 0 
                     THEN 'Reviewer 1' ELSE 'Reviewer 2' END
            ELSE 
                CASE (ROW_NUMBER() OVER (PARTITION BY is_high_weightage ORDER BY occurrence_count DESC, ProductName ASC) - 1) % 3
                    WHEN 0 THEN 'Reviewer 3'
                    WHEN 1 THEN 'Reviewer 4'
                    ELSE 'Reviewer 5'
                END
        END AS ProductReviewer,
        CASE WHEN is_high_weightage = 1 THEN 'High Weightage' ELSE 'Lower Weightage' END AS WeightageCategory
    FROM product_with_high_low
),

-- Step 8: Join articles with product assignments
article_with_products AS (
    SELECT 
        bd.ArticleId,
        bd.ProductId,
        bd.ProductName,
        pfa.ProductReviewer,
        pfa.occurrence_count,
        pfa.product_rank
    FROM base_data bd
    INNER JOIN product_final_assignment pfa ON bd.ProductId = pfa.ProductId
),

-- Step 9: Get primary product for each article
article_primary_product AS (
    SELECT 
        ArticleId,
        MIN(product_rank) AS min_rank
    FROM article_with_products
    GROUP BY ArticleId
),

-- Step 10: Determine article's assigned reviewer
article_reviewer AS (
    SELECT 
        awp.ArticleId,
        awp.ProductReviewer AS ArticleAssignedReviewer
    FROM article_with_products awp
    INNER JOIN article_primary_product app 
        ON awp.ArticleId = app.ArticleId 
        AND awp.product_rank = app.min_rank
)

-- Final output
SELECT 
    awp.ArticleId,
    awp.ProductName,
    ar.ArticleAssignedReviewer AS AssignedReviewer
FROM article_with_products awp
INNER JOIN article_reviewer ar ON awp.ArticleId = ar.ArticleId
WHERE awp.ProductReviewer = ar.ArticleAssignedReviewer
ORDER BY ar.ArticleAssignedReviewer, awp.ProductName, awp.ArticleId;
"""


# =============================================================================
# FUNCTIONS
# =============================================================================

def get_date_range():
    """Get the date range for the query (last Monday's data)."""
    today = datetime.now()
    # Find last Monday (or today if it's Monday)
    days_since_monday = today.weekday()
    last_monday = today - timedelta(days=days_since_monday)
    
    start_date = last_monday.strftime('%Y-%m-%d 00:00:00')
    end_date = last_monday.strftime('%Y-%m-%d 23:59:59')
    
    return start_date, end_date


def connect_to_database():
    """Establish connection to MySQL database."""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        if connection.is_connected():
            print(f"✓ Connected to MySQL database: {DB_CONFIG['database']}")
            return connection
    except Error as e:
        print(f"✗ Error connecting to MySQL: {e}")
        raise


def run_assignment_query(connection, start_date, end_date):
    """Execute the reviewer assignment query and return results as DataFrame."""
    try:
        print(f"Running query for date range: {start_date} to {end_date}")
        df = pd.read_sql(REVIEWER_ASSIGNMENT_QUERY, connection, params=(start_date, end_date))
        print(f"✓ Query executed successfully. {len(df)} rows returned.")
        return df
    except Error as e:
        print(f"✗ Error executing query: {e}")
        raise


def generate_excel_report(df, output_dir='output'):
    """Generate Excel report from the assignment data."""
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Reviewer_Assignments_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb = Workbook()
    
    # Sheet 1: Article-Product Assignments
    ws1 = wb.active
    ws1.title = "Article-Product Assignments"
    
    headers = ['ArticleId', 'ProductName', 'AssignedReviewer']
    ws1.append(headers)
    
    for _, row in df.iterrows():
        ws1.append([row['ArticleId'], row['ProductName'], row['AssignedReviewer']])
    
    # Sheet 2: Reviewer Summary
    ws2 = wb.create_sheet("Reviewer Summary")
    ws2.append(['Reviewer', 'TotalArticles', 'TotalProducts'])
    
    summary = df.groupby('AssignedReviewer').agg({
        'ArticleId': 'nunique',
        'ProductName': 'nunique'
    }).reset_index()
    
    for _, row in summary.iterrows():
        ws2.append([row['AssignedReviewer'], row['ArticleId'], row['ProductName']])
    
    # Formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    wb.save(filepath)
    print(f"✓ Excel report saved: {filepath}")
    return filepath


def send_email(to_email, reviewer_name, df_reviewer, report_filepath, date_range):
    """Send email notification to a reviewer."""
    msg = MIMEMultipart()
    msg['From'] = EMAIL_CONFIG['from_email']
    msg['To'] = to_email
    msg['Subject'] = f"Literature Review Assignment - {reviewer_name} - {datetime.now().strftime('%d %b %Y')}"
    
    # Email body
    article_count = df_reviewer['ArticleId'].nunique()
    product_count = df_reviewer['ProductName'].nunique()
    product_list = df_reviewer['ProductName'].unique()
    
    body = f"""
    <html>
    <body>
    <p>Dear {reviewer_name},</p>
    
    <p>Your weekly literature review assignment is ready.</p>
    
    <h3>Assignment Summary:</h3>
    <ul>
        <li><strong>Total Articles:</strong> {article_count}</li>
        <li><strong>Total Products:</strong> {product_count}</li>
        <li><strong>Products Assigned:</strong> {', '.join(product_list)}</li>
        <li><strong>Data Date:</strong> {date_range[0][:10]}</li>
    </ul>
    
    <p>Please find the detailed assignment in the attached Excel file.</p>
    
    <p>Best regards,<br>
    Automated Assignment System</p>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(body, 'html'))
    
    # Attach Excel file
    with open(report_filepath, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(report_filepath)}"')
        msg.attach(part)
    
    # Send email
    try:
        with smtplib.SMTP(EMAIL_CONFIG['smtp_host'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['smtp_user'], EMAIL_CONFIG['smtp_password'])
            server.sendmail(EMAIL_CONFIG['from_email'], to_email, msg.as_string())
        print(f"✓ Email sent to {reviewer_name} ({to_email})")
    except Exception as e:
        print(f"✗ Failed to send email to {reviewer_name}: {e}")


# def send_notifications(df, report_filepath, date_range):
#     """Send email notifications to all reviewers."""
#     print("\n--- Sending Email Notifications ---")
    
#     for reviewer, email in REVIEWER_EMAILS.items():
#         if email:
#             df_reviewer = df[df['AssignedReviewer'] == reviewer]
#             if not df_reviewer.empty:
#                 send_email(email, reviewer, df_reviewer, report_filepath, date_range)
#             else:
#                 print(f"⚠ No assignments for {reviewer}, skipping email.")
#         else:
#             print(f"⚠ No email configured for {reviewer}")

def send_notifications(df, report_filepath, date_range):
    """Send email notification with full report to the configured Gmail."""
    print("\n--- Sending Email Notification ---")
    
    to_email = 'pritam.kundilya@kretacx.com'
    
    msg = MIMEMultipart()
    msg['From'] = EMAIL_CONFIG['from_email']
    msg['To'] = to_email
    msg['Subject'] = f"Literature Review Assignments - {datetime.now().strftime('%d %b %Y')}"
    
    # Build summary for all reviewers
    summary_lines = []
    for reviewer in df['AssignedReviewer'].unique():
        df_reviewer = df[df['AssignedReviewer'] == reviewer]
        article_count = df_reviewer['ArticleId'].nunique()
        product_count = df_reviewer['ProductName'].nunique()
        summary_lines.append(f"<li><strong>{reviewer}:</strong> {article_count} articles, {product_count} products</li>")
    
    body = f"""
    <html>
    <body>
    <p>Hello,</p>
    
    <p>The weekly literature review assignments are ready.</p>
    
    <h3>Assignment Summary:</h3>
    <ul>
        {''.join(summary_lines)}
    </ul>
    <p><strong>Data Date:</strong> {date_range[0][:10]}</p>
    
    <p>Please find the detailed assignment in the attached Excel file.</p>
    
    <p>Best regards,<br>
    Automated Assignment System</p>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(body, 'html'))
    
    # Attach Excel file
    with open(report_filepath, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(report_filepath)}"')
        msg.attach(part)
    
    # Send via Gmail SMTP
    try:
        with smtplib.SMTP(EMAIL_CONFIG['smtp_host'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['smtp_user'], EMAIL_CONFIG['smtp_password'])
            server.sendmail(EMAIL_CONFIG['from_email'], to_email, msg.as_string())
        print(f"✓ Email sent to {to_email}")
    except Exception as e:
        print(f"✗ Failed to send email: {e}")

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    print("=" * 60)
    print("REVIEWER ASSIGNMENT AUTOMATION")
    print(f"Run Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # Get date range
    start_date, end_date = get_date_range()
    print(f"\nDate Range: {start_date} to {end_date}")
    
    # Connect to database
    connection = connect_to_database()
    
    try:
        # Run the assignment query
        df = run_assignment_query(connection, start_date, end_date)
        
        if df.empty:
            print("⚠ No data found for the specified date range.")
            return
        
        # Print summary
        print("\n--- Assignment Summary ---")
        summary = df.groupby('AssignedReviewer').agg({
            'ArticleId': 'nunique',
            'ProductName': 'nunique'
        })
        print(summary)
        
        # Generate Excel report
        report_filepath = generate_excel_report(df)
        
        # Send email notifications
        send_notifications(df, report_filepath, (start_date, end_date))
        
        print("\n" + "=" * 60)
        print("✓ AUTOMATION COMPLETED SUCCESSFULLY")
        print("=" * 60)
        
    finally:
        if connection.is_connected():
            connection.close()
            print("Database connection closed.")


if __name__ == "__main__":
    main()
